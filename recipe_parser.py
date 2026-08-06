# -*- coding: utf-8 -*-
"""
recipe_parser.py
หน้าที่: อ่านไฟล์ Recipe Excel (Food / Beverage / Dessert) แล้วดึงรายการ ingredient ออกมา
ในรูปแบบเดียวกันหมด ไม่ว่าไฟล์ต้นทางจะมี layout ต่างกันแค่ไหน

รู้จัก 3 ประเภทชีท:
  1) INGREDIENTS  - ชีท "Ingredients" : master pack size (PU/RU/#RU per PU) ของแต่ละ RM/WIP
  2) FG_RECIPE    - ชีทสูตรเมนูขาย (Recipe_xxx, เสร็จ-Recipe xxx) : FG code + ingredient rows
  3) WIP_RECIPE   - ชีทสูตรผลิต WIP จาก RM (xxx_Branch, WIP_xxx) : WIP code + BATCH + ingredient rows
ชีทอื่น (item master dump, cost summary ล้วน) จะถูกข้ามอัตโนมัติ เพราะไม่มี header ที่รู้จัก
"""
import openpyxl
from unit_utils import normalize_unit

FG_HEADERS = {"cm-pos", "menu", "item code", "qty", "small unit"}
WIP_HEADERS = {"batch", "recipe name", "item code", "qty", "small unit"}
ING_HEADERS = {"itemcode", "pu u/m", "ru u/m", "# ru per pu"}


def _norm_header(text):
    """lowercase + ยุบช่องว่างซ้ำให้เหลือช่องเดียว (กันกรณี '#  RU per PU' มี 2 ช่องว่าง)"""
    return " ".join(str(text).strip().lower().split())


def _row_text_set(row, limit=14):
    out = set()
    for c in row[:limit]:
        if isinstance(c, str):
            out.add(_norm_header(c))
    return out


def _find_header_row(ws, max_scan=8):
    """สแกน 8 แถวแรกของชีท หา header row ที่ตรงกับ pattern ที่รู้จัก คืน (row_idx(1-based), header_list, kind)"""
    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for idx, row in enumerate(rows, start=1):
        if row is None:
            continue
        txt = _row_text_set(row)
        if ING_HEADERS.issubset(txt):
            return idx, row, "INGREDIENTS"
        if WIP_HEADERS.issubset(txt):
            return idx, row, "WIP_RECIPE"
        if FG_HEADERS.issubset(txt):
            return idx, row, "FG_RECIPE"
    return None, None, "SKIP"


def _col_index(header_row, *names):
    """หา index (0-based) ของคอลัมน์ตามชื่อ (case-insensitive) คืน None ถ้าไม่เจอ
    ถ้ามีชื่อซ้ำกันหลายคอลัมน์ (เช่น 'Item Code' โผล่ 2 รอบ) คืน "ตัวแรก" เว้นแต่ระบุ occurrence"""
    lower = [_norm_header(h) if h else "" for h in header_row]
    for name in names:
        if name in lower:
            return lower.index(name)
    return None


def _col_indices_all(header_row, name):
    lower = [_norm_header(h) if h else "" for h in header_row]
    return [i for i, h in enumerate(lower) if h == name]


def parse_ingredients_sheet(ws, header_row_idx, header_row):
    """คืน dict: item_code -> {pu_unit, ru_unit, ru_per_pu, yield_pct, case_desc}"""
    idx_code = _col_index(header_row, "itemcode")
    idx_case = _col_index(header_row, "case")
    idx_pu = _col_index(header_row, "pu u/m")
    idx_ru = _col_index(header_row, "ru u/m")
    idx_qty = _col_index(header_row, "#  ru per pu", "# ru per pu")
    idx_yield = _col_index(header_row, "yield %")
    out = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or idx_code is None:
            continue
        code = row[idx_code] if idx_code < len(row) else None
        if not code:
            continue
        code = str(code).strip().upper()
        try:
            ru_per_pu = float(row[idx_qty]) if idx_qty is not None and row[idx_qty] not in (None, "") else None
        except (TypeError, ValueError):
            ru_per_pu = None
        try:
            yield_pct = float(row[idx_yield]) if idx_yield is not None and row[idx_yield] not in (None, "") else 1.0
        except (TypeError, ValueError):
            yield_pct = 1.0
        out[code] = {
            "case_desc": row[idx_case] if idx_case is not None and idx_case < len(row) else None,
            "pu_unit": row[idx_pu] if idx_pu is not None and idx_pu < len(row) else None,
            "ru_unit": row[idx_ru] if idx_ru is not None and idx_ru < len(row) else None,
            "ru_per_pu": ru_per_pu,
            "yield_pct": yield_pct or 1.0,
        }
    return out


def parse_fg_recipe_sheet(ws, header_row_idx, header_row, source_file, sheet_name):
    """สูตรเมนูขาย (FG) — คืน list ของ dict ingredient row
    รูปแบบคอลัมน์: No | CM-POS | Total Cost | Menu | Item Code | Item Description |
                   QTY | Small Unit | Unit Cost | Cost | Remark | [QTY ต้นฉบับ]"""
    idx_fg_code = _col_index(header_row, "cm-pos")
    idx_menu = _col_index(header_row, "menu")
    idx_ing_code = _col_index(header_row, "item code")
    idx_ing_desc = _col_index(header_row, "item description")
    idx_qty = _col_index(header_row, "qty")
    idx_unit = _col_index(header_row, "small unit")
    idx_remark = _col_index(header_row, "remark")
    # QTY อาจโผล่ 2 รอบ (ตัวหลัง = ค่าต้นฉบับก่อนแปลง)
    qty_positions = _col_indices_all(header_row, "qty")
    idx_orig_qty = qty_positions[1] if len(qty_positions) > 1 else None

    rows_out = []
    cur_fg_code, cur_menu = None, None
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row:
            continue
        # แถวที่มี fg code ใหม่ = เริ่มเมนูใหม่
        fg_code_cell = row[idx_fg_code] if idx_fg_code is not None and idx_fg_code < len(row) else None
        if fg_code_cell:
            cur_fg_code = str(fg_code_cell).strip().upper()
            cur_menu = row[idx_menu] if idx_menu is not None and idx_menu < len(row) else None

        ing_code = row[idx_ing_code] if idx_ing_code is not None and idx_ing_code < len(row) else None
        if not ing_code or not cur_fg_code:
            continue
        qty_val = row[idx_qty] if idx_qty is not None and idx_qty < len(row) else None
        if qty_val in (None, ""):
            continue
        try:
            qty_val = float(qty_val)
        except (TypeError, ValueError):
            continue

        orig_qty = None
        if idx_orig_qty is not None and idx_orig_qty < len(row) and row[idx_orig_qty] not in (None, ""):
            try:
                orig_qty = float(row[idx_orig_qty])
            except (TypeError, ValueError):
                orig_qty = None

        raw_unit = row[idx_unit] if idx_unit is not None and idx_unit < len(row) else None
        rows_out.append({
            "source_file": source_file,
            "sheet": sheet_name,
            "recipe_type": "FG",
            "parent_code": cur_fg_code,
            "parent_name": cur_menu,
            "batch_qty": None,
            "batch_unit": None,
            "ingredient_code": str(ing_code).strip().upper(),
            "ingredient_desc": row[idx_ing_desc] if idx_ing_desc is not None and idx_ing_desc < len(row) else None,
            "excel_qty": qty_val,               # ค่าที่ไฟล์เตรียมไว้ให้กรอกใน CM POS (อาจแปลงแล้วหรือยังไม่แปลงก็ได้)
            "excel_unit_raw": raw_unit,
            "excel_unit": normalize_unit(raw_unit),
            "excel_qty_original": orig_qty,      # ค่าต้นฉบับก่อนแปลง (ถ้าไฟล์มีคอลัมน์นี้)
            "remark": row[idx_remark] if idx_remark is not None and idx_remark < len(row) else None,
        })
    return rows_out


def parse_wip_recipe_sheet(ws, header_row_idx, header_row, source_file, sheet_name):
    """สูตรผลิต WIP จาก RM — คอลัมน์: No | Item Code(WIP) | Total Cost | Recipe Name | BATCH | Unit |
                                        Item Code(RM) | Item Description | QTY | Small Unit | Unit Cost | Cost | Remark"""
    wip_code_positions = _col_indices_all(header_row, "item code")
    idx_wip_code = wip_code_positions[0] if wip_code_positions else None
    idx_ing_code = wip_code_positions[1] if len(wip_code_positions) > 1 else None
    idx_recipe_name = _col_index(header_row, "recipe name")
    idx_batch = _col_index(header_row, "batch")
    idx_batch_unit = None
    unit_positions = _col_indices_all(header_row, "unit")
    if unit_positions:
        idx_batch_unit = unit_positions[0]
    idx_ing_desc = _col_index(header_row, "item description")
    idx_qty = _col_index(header_row, "qty")
    idx_small_unit = _col_index(header_row, "small unit")
    idx_remark = _col_index(header_row, "remark")

    rows_out = []
    cur_wip_code, cur_name, cur_batch, cur_batch_unit = None, None, None, None
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row:
            continue
        wip_cell = row[idx_wip_code] if idx_wip_code is not None and idx_wip_code < len(row) else None
        if wip_cell:
            cur_wip_code = str(wip_cell).strip().upper()
            cur_name = row[idx_recipe_name] if idx_recipe_name is not None and idx_recipe_name < len(row) else None
            cur_batch = row[idx_batch] if idx_batch is not None and idx_batch < len(row) else None
            cur_batch_unit = row[idx_batch_unit] if idx_batch_unit is not None and idx_batch_unit < len(row) else None

        ing_code = row[idx_ing_code] if idx_ing_code is not None and idx_ing_code < len(row) else None
        if not ing_code or not cur_wip_code:
            continue
        qty_val = row[idx_qty] if idx_qty is not None and idx_qty < len(row) else None
        if qty_val in (None, ""):
            continue
        try:
            qty_val = float(qty_val)
        except (TypeError, ValueError):
            continue

        raw_unit = row[idx_small_unit] if idx_small_unit is not None and idx_small_unit < len(row) else None
        rows_out.append({
            "source_file": source_file,
            "sheet": sheet_name,
            "recipe_type": "WIP",
            "parent_code": cur_wip_code,
            "parent_name": cur_name,
            "batch_qty": cur_batch,
            "batch_unit": cur_batch_unit,
            "ingredient_code": str(ing_code).strip().upper(),
            "ingredient_desc": row[idx_ing_desc] if idx_ing_desc is not None and idx_ing_desc < len(row) else None,
            "excel_qty": qty_val,
            "excel_unit_raw": raw_unit,
            "excel_unit": normalize_unit(raw_unit),
            "excel_qty_original": None,
            "remark": row[idx_remark] if idx_remark is not None and idx_remark < len(row) else None,
        })
    return rows_out


def parse_workbook(path, source_file_label):
    """อ่าน 1 ไฟล์ Excel ทั้งไฟล์ คืน (ingredient_rows, ingredients_lookup)"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ingredient_rows = []
    ingredients_lookup = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_idx, header_row, kind = _find_header_row(ws)
        if kind == "SKIP" or header_idx is None:
            continue
        if kind == "INGREDIENTS":
            ingredients_lookup.update(parse_ingredients_sheet(ws, header_idx, header_row))
        elif kind == "FG_RECIPE":
            ingredient_rows.extend(parse_fg_recipe_sheet(ws, header_idx, header_row, source_file_label, sheet_name))
        elif kind == "WIP_RECIPE":
            ingredient_rows.extend(parse_wip_recipe_sheet(ws, header_idx, header_row, source_file_label, sheet_name))
    wb.close()
    return ingredient_rows, ingredients_lookup


def parse_all(files):
    """files: list ของ (path, label) เช่น [('food.xlsx','Food'), ...]
    คืน (all_ingredient_rows, merged_ingredients_lookup)"""
    all_rows = []
    merged_lookup = {}
    for path, label in files:
        rows, lookup = parse_workbook(path, label)
        all_rows.extend(rows)
        # อย่า overwrite ถ้ามีอยู่แล้วจากไฟล์อื่น (กันข้ามไฟล์ปนกัน) — เก็บทุกไฟล์ไว้แยก key ก็ได้ แต่ทำ merge ง่ายๆ ก่อน
        for k, v in lookup.items():
            if k not in merged_lookup:
                merged_lookup[k] = v
    return all_rows, merged_lookup
